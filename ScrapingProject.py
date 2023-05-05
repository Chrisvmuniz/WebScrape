# Twilio credentials

import time
from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
from twilio.rest import Client
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


#==================[Twilio Stuff]===================#
account_sid = "ACc19220d900396bfb659936740c4270a2"  #
auth_token = "d2d0147deda3e932ac7c759363d5a952"     #
twilio_number = '+18777604420'                          #
my_number = '+18324897727'                              #
#===================================================#

#Website URL I am scraping
url = 'https://crypto.com/price'  #Originally used https://coinmarketcap.com/ but had issues
#Headers
request_headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
#I make my client here because I don't want to remake it in a loop or anything and can access it anywhere
client = Client(account_sid, auth_token)

#Start with no previous Prices
previous_prices = {'BTC': None, 'ETH': None}
#Start with a count of zero
count = 0

#Loop so we can contiunually check the price
while True:
    #Add one to our run count each loop
    count = count + 1
    #Print the run count to console so we can visually see the program is running
    print("Run " + str(count))
    #Request the url and header
    req = Request(url, headers=request_headers)
    #Open the website
    webpage = urlopen(req).read()
    #Use BeautifulSoup to parse
    soup = BeautifulSoup(webpage, 'html.parser')
    #Get the table rows
    table_rows = soup.find('table', class_='chakra-table css-1qpk7f7')

    #Now we make our workboot
    workbook = openpyxl.Workbook()
    #Set the active worksheet
    worksheet = workbook.active

    #Format the worksheets header and put a color to be fancy
    header_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')

    #Now we title the headers for the worksheet and loop for format
    headers = ['Name', 'Symbol', 'Current Price', '% Change (24h)', 'Price Change (24h)']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    #Now we set our row data
    for row_num, record in enumerate(table_rows.find_all('tbody'), 2):
        rows = record.find_all('tr')
        for i, row in enumerate(rows[:5]):
            #Set everything
            td = row.findAll('td')
            name = row.find('p', class_='chakra-text css-rkws3').text
            symbol = row.find('span', class_='chakra-text css-1jj7b1a').text
            price = float(row.find('div', class_='css-b1ilzc').text.replace('$', '').replace(',', ''))
            change = float(td[4].text.replace('%', ''))
            cor_price = round(price * ((100.00 - change) / 100.00), 2)

            #Now we save our data
            values = [name, symbol, f"${price:.2f}", f"{change}%", f"${cor_price:.2f}"]
            for col_num, value in enumerate(values, 1):
                cell = worksheet.cell(row=row_num + i, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            #Now we send an alert if Bitcoin or Etherium is lower or higher (we can easily do this by taking abs value to tell the change)
            if symbol in previous_prices:
                #if we have a change on one of our two crypto's then we notify
                if previous_prices[symbol] is not None and (abs(previous_prices[symbol] - price) >= 5):
                    if symbol == 'BTC':
                        change = previous_prices[symbol] - price
                        textmsg = client.messages.create(to=my_number,from_=twilio_number,body="Bitcoin price changed by $"+str(change))
                        #Print to console to make sure it sent to right sid
                        print(textmsg.sid)

                    elif symbol == 'ETH':
                        change = previous_prices[symbol] - price
                        textmsg = client.messages.create(to=my_number,from_=twilio_number,body="Bitcoin price changed by $"+str(change))
                        #Print to console to make sure it sent to right sid
                        print(textmsg.sid)
            #I just had this for testing to make sure there was no change/That my twilio was working
            """
            else:
                textmsg = client.messages.create(to=my_number,from_=twilio_number,body="No Price Change Yet")
                print(textmsg.sid)
            previous_prices[symbol] = price
            """

    #Now Save the Excel Spreadsheet, I have it to save a new version file every time in case we want to see history
    workbook.save(f'Crypto_Report{count}.xlsx')

    # Pause for (500 seconds) before starting the next iteration
    time.sleep(500)
